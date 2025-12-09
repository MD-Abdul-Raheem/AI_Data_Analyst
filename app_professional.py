import os
import json
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import base64

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

class ProfessionalDataAnalyst:
    def __init__(self, df):
        self.df = df.copy()
        self.original_df = df.copy()
        
    def analyze(self):
        """Complete professional analysis"""
        results = {}
        
        # 1. Data Overview
        results['overview'] = {
            'total_records': len(self.df),
            'total_columns': len(self.df.columns),
            'columns': list(self.df.columns),
            'dtypes': {str(k): str(v) for k, v in self.df.dtypes.items()}
        }
        
        # 2. Clean Data
        self.df = self.df.dropna(how='all')
        self.df = self.df.drop_duplicates()
        
        # 3. Statistical Summary
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        results['statistics'] = {}
        
        for col in numeric_cols:
            results['statistics'][col] = {
                'mean': float(self.df[col].mean()),
                'median': float(self.df[col].median()),
                'std': float(self.df[col].std()),
                'min': float(self.df[col].min()),
                'max': float(self.df[col].max()),
                'range': float(self.df[col].max() - self.df[col].min())
            }
        
        # 4. Categorical Summary
        cat_cols = self.df.select_dtypes(include=['object']).columns.tolist()
        results['categories'] = {}
        
        for col in cat_cols[:5]:
            top_5 = self.df[col].value_counts().head(5)
            results['categories'][col] = {str(k): int(v) for k, v in top_5.items()}
        
        # 5. Correlations
        if len(numeric_cols) > 1:
            corr = self.df[numeric_cols].corr()
            results['correlations'] = {
                str(i): {str(j): float(corr.loc[i, j]) for j in corr.columns}
                for i in corr.index
            }
        
        # 6. Business Insights
        results['insights'] = []
        results['insights'].append(f"Dataset contains {len(self.df):,} records and {len(self.df.columns)} columns")
        
        for col in numeric_cols[:3]:
            total = self.df[col].sum()
            avg = self.df[col].mean()
            results['insights'].append(f"{col}: Total = {total:,.2f}, Average = {avg:,.2f}")
        
        return results
    
    def create_excel_report(self):
        """Create professional Excel report"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Cleaned Data
        ws1 = wb.create_sheet("Cleaned_Data")
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws1.append(r)
        
        # Format header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sheet 2: Statistics
        ws2 = wb.create_sheet("Statistics")
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        if len(numeric_cols) > 0:
            stats = self.df[numeric_cols].describe()
            stats.loc['range'] = stats.loc['max'] - stats.loc['min']
            
            for r in dataframe_to_rows(stats, index=True, header=True):
                ws2.append(r)
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Sheet 3: Summary Metrics
        ws3 = wb.create_sheet("Summary_Metrics")
        ws3.append(["Metric", "Value"])
        ws3.cell(1, 1).font = Font(bold=True, color="FFFFFF")
        ws3.cell(1, 2).font = Font(bold=True, color="FFFFFF")
        ws3.cell(1, 1).fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        ws3.cell(1, 2).fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        
        ws3.append(["Total Records", len(self.df)])
        ws3.append(["Total Columns", len(self.df.columns)])
        ws3.append(["Numeric Columns", len(numeric_cols)])
        ws3.append(["Missing Values", int(self.df.isnull().sum().sum())])
        
        # Add KPI metrics
        for col in numeric_cols:
            if any(kw in col.lower() for kw in ['sales', 'profit', 'revenue', 'amount']):
                ws3.append([f"Total {col}", round(float(self.df[col].sum()), 2)])
                ws3.append([f"Average {col}", round(float(self.df[col].mean()), 2)])
                ws3.append([f"Median {col}", round(float(self.df[col].median()), 2)])
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        return wb

@app.route('/')
def index():
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Professional Data Analyst</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; }
            .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
            h1 { color: #2c3e50; margin-bottom: 30px; }
            .upload-area { border: 2px dashed #3498db; padding: 40px; text-align: center; border-radius: 10px; margin-bottom: 30px; cursor: pointer; }
            .upload-area:hover { background: #ecf0f1; }
            .btn { background: #3498db; color: white; padding: 12px 30px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; margin: 10px; }
            .btn:hover { background: #2980b9; }
            .results { margin-top: 30px; display: none; }
            .section { margin: 20px 0; padding: 20px; background: #f8f9fa; border-radius: 5px; }
            .section h2 { color: #2c3e50; margin-bottom: 15px; }
            .metric { display: inline-block; margin: 10px; padding: 15px 25px; background: #3498db; color: white; border-radius: 5px; }
            table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
            th { background: #3498db; color: white; }
            .loading { display: none; text-align: center; padding: 40px; }
            .spinner { border: 4px solid #f3f3f3; border-top: 4px solid #3498db; border-radius: 50%; width: 50px; height: 50px; animation: spin 1s linear infinite; margin: 0 auto; }
            @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🎯 Professional Data Analyst</h1>
            
            <div class="upload-area" onclick="document.getElementById('file').click()">
                <h2>📊 Upload Your Dataset</h2>
                <p>Click or drag CSV/Excel file here</p>
                <input type="file" id="file" style="display:none" accept=".csv,.xlsx,.xls">
            </div>
            
            <div class="loading" id="loading">
                <div class="spinner"></div>
                <p>Analyzing your data...</p>
            </div>
            
            <div class="results" id="results">
                <div class="section">
                    <h2>📈 Overview</h2>
                    <div id="overview"></div>
                </div>
                
                <div class="section">
                    <h2>📊 Statistics</h2>
                    <div id="statistics"></div>
                </div>
                
                <div class="section">
                    <h2>💡 Key Insights</h2>
                    <div id="insights"></div>
                </div>
                
                <div class="section">
                    <h2>📥 Downloads</h2>
                    <button class="btn" onclick="downloadExcel()">📊 Download Excel Report</button>
                </div>
            </div>
        </div>
        
        <script>
            document.getElementById('file').addEventListener('change', async (e) => {
                const file = e.target.files[0];
                if (!file) return;
                
                const formData = new FormData();
                formData.append('file', file);
                
                document.getElementById('loading').style.display = 'block';
                document.getElementById('results').style.display = 'none';
                
                try {
                    const response = await fetch('/analyze', {
                        method: 'POST',
                        body: formData
                    });
                    
                    const data = await response.json();
                    
                    if (data.error) {
                        alert('Error: ' + data.error);
                        return;
                    }
                    
                    displayResults(data);
                } catch (error) {
                    alert('Error: ' + error.message);
                } finally {
                    document.getElementById('loading').style.display = 'none';
                }
            });
            
            function displayResults(data) {
                // Overview
                document.getElementById('overview').innerHTML = `
                    <div class="metric">Records: ${data.overview.total_records.toLocaleString()}</div>
                    <div class="metric">Columns: ${data.overview.total_columns}</div>
                `;
                
                // Statistics
                let statsHTML = '<table><tr><th>Column</th><th>Mean</th><th>Median</th><th>Min</th><th>Max</th><th>Range</th></tr>';
                for (const [col, stats] of Object.entries(data.statistics)) {
                    statsHTML += `<tr>
                        <td><strong>${col}</strong></td>
                        <td>${stats.mean.toFixed(2)}</td>
                        <td>${stats.median.toFixed(2)}</td>
                        <td>${stats.min.toFixed(2)}</td>
                        <td>${stats.max.toFixed(2)}</td>
                        <td>${stats.range.toFixed(2)}</td>
                    </tr>`;
                }
                statsHTML += '</table>';
                document.getElementById('statistics').innerHTML = statsHTML;
                
                // Insights
                let insightsHTML = '<ul>';
                data.insights.forEach(insight => {
                    insightsHTML += `<li style="margin: 10px 0; font-size: 16px;">✓ ${insight}</li>`;
                });
                insightsHTML += '</ul>';
                document.getElementById('insights').innerHTML = insightsHTML;
                
                document.getElementById('results').style.display = 'block';
            }
            
            function downloadExcel() {
                window.location.href = '/download/excel';
            }
        </script>
    </body>
    </html>
    '''

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Load data
        if filename.endswith('.csv'):
            df = pd.read_csv(filepath)
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filepath)
        else:
            return jsonify({'error': 'Unsupported file format'}), 400
        
        # Analyze
        analyst = ProfessionalDataAnalyst(df)
        results = analyst.analyze()
        
        # Create Excel
        wb = analyst.create_excel_report()
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Professional_Analysis_Report.xlsx')
        wb.save(excel_path)
        
        return jsonify(results)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/excel')
def download_excel():
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Professional_Analysis_Report.xlsx')
    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True, download_name='Professional_Analysis_Report.xlsx')
    return jsonify({'error': 'File not found'}), 404

if __name__ == '__main__':
    print("=" * 60)
    print("PROFESSIONAL DATA ANALYST")
    print("=" * 60)
    print("Server starting on http://localhost:5000")
    print("Upload CSV or Excel files")
    print("Get instant professional analysis")
    print("Download Excel report with all metrics")
    print("=" * 60)
    app.run(debug=True, port=5000)
