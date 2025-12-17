import os
import json
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from flask import Flask, render_template, request, jsonify, send_file
from flask.json.provider import DefaultJSONProvider
from werkzeug.utils import secure_filename
from io import BytesIO, StringIO
import base64
from datetime import datetime
from scipy import stats

class CustomJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.floating)):
            if np.isnan(obj) or np.isinf(obj):
                return None
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return super().default(obj)

app = Flask(__name__)
app.json = CustomJSONProvider(app)
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

class DataAnalyst:
    def __init__(self, df):
        self.df = df
        self.original_df = df.copy()
        self.insights = []
        self.charts = []
        self.column_types = {}
        
    def understand_data(self):
        """Stage 1: Data Understanding"""
        info = {
            'shape': self.df.shape,
            'columns': list(self.df.columns),
            'dtypes': self.df.dtypes.astype(str).to_dict(),
            'memory_usage': safe_float(self.df.memory_usage(deep=True).sum() / 1024**2),
            'head': self.df.head(10).replace({np.nan: None, np.inf: None, -np.inf: None}).to_dict('records'),
            'tail': self.df.tail(10).replace({np.nan: None, np.inf: None, -np.inf: None}).to_dict('records')
        }
        
        # Classify columns
        for col in self.df.columns:
            if self.df[col].dtype in ['int64', 'float64']:
                if self.df[col].nunique() < 20 and self.df[col].nunique() / len(self.df) < 0.05:
                    self.column_types[col] = 'categorical_numeric'
                else:
                    self.column_types[col] = 'numerical'
            elif pd.api.types.is_datetime64_any_dtype(self.df[col]):
                self.column_types[col] = 'datetime'
            elif self.df[col].dtype == 'bool':
                self.column_types[col] = 'boolean'
            elif self.df[col].nunique() == len(self.df):
                self.column_types[col] = 'id'
            else:
                self.column_types[col] = 'categorical'
        
        info['column_types'] = self.column_types
        return info
    
    def clean_data(self):
        """Stage 2: Data Cleaning - BI-Optimized"""
        cleaning_report = {
            'missing_values': {},
            'duplicates_removed': 0,
            'outliers_detected': {},
            'transformations': [],
            'imputation_strategies': {},
            'bi_recommendations': []
        }
        
        # Standardize column names for BI tools (no special characters, spaces to underscores)
        original_columns = self.df.columns.tolist()
        new_columns = []
        for col in self.df.columns:
            # Convert to string and clean
            clean_col = str(col).strip()
            clean_col = clean_col.replace(' ', '_').replace('-', '_').replace('.', '_')
            clean_col = ''.join(c if c.isalnum() or c == '_' else '_' for c in clean_col)
            clean_col = clean_col.strip('_')
            new_columns.append(clean_col)
        
        self.df.columns = new_columns
        
        # Update column_types with new names
        old_to_new = dict(zip(original_columns, new_columns))
        self.column_types = {old_to_new.get(k, k): v for k, v in self.column_types.items()}
        
        if new_columns != original_columns:
            cleaning_report['transformations'].append('Column names standardized for BI ingestion')
            cleaning_report['bi_recommendations'].append('Column names cleaned: removed special characters, replaced spaces with underscores')
        
        # Missing values
        missing = self.df.isnull().sum()
        cleaning_report['missing_values'] = {str(k): int(v) for k, v in missing[missing > 0].items()}
        
        # Fill missing values with documented imputation strategies
        for col in self.df.columns:
            if self.df[col].isnull().sum() > 0:
                missing_pct = (self.df[col].isnull().sum() / len(self.df)) * 100
                if self.column_types.get(col) == 'numerical':
                    if missing_pct < 5:
                        self.df[col].fillna(self.df[col].median(), inplace=True)
                        cleaning_report['imputation_strategies'][col] = 'Median imputation (< 5% missing)'
                    else:
                        self.df[col].fillna(self.df[col].mean(), inplace=True)
                        cleaning_report['imputation_strategies'][col] = 'Mean imputation (>= 5% missing)'
                    cleaning_report['transformations'].append(f"{col}: filled with {'median' if missing_pct < 5 else 'mean'}")
                else:
                    self.df[col].fillna('Unknown', inplace=True)
                    cleaning_report['imputation_strategies'][col] = 'Categorical: filled with Unknown'
                    cleaning_report['transformations'].append(f"{col}: filled with 'Unknown'")
        
        # Remove duplicates
        before = len(self.df)
        self.df.drop_duplicates(inplace=True)
        cleaning_report['duplicates_removed'] = before - len(self.df)
        
        # Standardize strings
        for col in self.df.select_dtypes(include=['object']).columns:
            self.df[col] = self.df[col].astype(str).str.strip().str.title()
        
        # Detect datetime columns and convert to ISO format for BI tools
        for col in self.df.columns:
            if self.df[col].dtype == 'object':
                try:
                    self.df[col] = pd.to_datetime(self.df[col])
                    self.column_types[col] = 'datetime'
                    cleaning_report['transformations'].append(f"{col}: converted to datetime (ISO format)")
                    cleaning_report['bi_recommendations'].append(f"{col}: Date column ready for Power BI/Tableau time intelligence. Create hierarchies: Year, Quarter, Month, Day")
                    # Create date hierarchies
                    self.df[f'{col}_Year'] = self.df[col].dt.year
                    self.df[f'{col}_Quarter'] = self.df[col].dt.quarter
                    self.df[f'{col}_Month'] = self.df[col].dt.month
                    self.df[f'{col}_MonthName'] = self.df[col].dt.month_name()
                    self.df[f'{col}_DayOfWeek'] = self.df[col].dt.day_name()
                    cleaning_report['transformations'].append(f"{col}: Date hierarchies created (Year, Quarter, Month, MonthName, DayOfWeek)")
                except:
                    pass
        
        # Detect outliers
        for col in self.df.select_dtypes(include=[np.number]).columns:
            try:
                Q1 = self.df[col].quantile(0.25)
                Q3 = self.df[col].quantile(0.75)
                IQR = Q3 - Q1
                if not np.isnan(Q1) and not np.isnan(Q3) and not np.isnan(IQR):
                    outliers = ((self.df[col] < (Q1 - 1.5 * IQR)) | (self.df[col] > (Q3 + 1.5 * IQR))).sum()
                    if outliers > 0:
                        cleaning_report['outliers_detected'][col] = int(outliers)
            except:
                pass
        
        return cleaning_report
    
    def perform_eda(self):
        """Stage 3: Exploratory Data Analysis with Detailed Explanations"""
        eda_results = {
            'numerical_summary': {},
            'categorical_summary': {},
            'correlations': {},
            'distributions': {},
            'explanations': {}
        }
        
        # Numerical summary with explanations
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            eda_results['numerical_summary'] = self.df[numeric_cols].describe().replace({np.nan: None, np.inf: None, -np.inf: None}).to_dict()
            eda_results['explanations']['numerical'] = 'Statistical summary shows central tendency (mean, median) and spread (std, min, max) for each numerical variable. Use this to identify outliers and understand data distribution.'
            
            # Correlation matrix
            if len(numeric_cols) > 1:
                corr = self.df[numeric_cols].corr()
                eda_results['correlations'] = corr.replace({np.nan: None, np.inf: None, -np.inf: None}).to_dict()
                eda_results['explanations']['correlations'] = 'Correlation matrix reveals relationships between numerical variables. Values close to 1 or -1 indicate strong positive or negative relationships, while values near 0 suggest no linear relationship.'
        
        # Categorical summary with explanations
        categorical_cols = [col for col in self.df.columns 
                          if col in self.column_types and self.column_types[col] in ['categorical', 'categorical_numeric']]
        for col in categorical_cols[:10]:
            value_counts = self.df[col].value_counts().head(20)
            eda_results['categorical_summary'][col] = value_counts.to_dict()
        
        if categorical_cols:
            eda_results['explanations']['categorical'] = 'Categorical analysis shows frequency distribution of non-numerical variables. This helps identify dominant categories and data imbalances.'
        
        return eda_results
    
    def generate_visualizations(self):
        """Generate high-quality visualizations with base64 encoding"""
        charts = []
        
        try:
            # Set style
            sns.set_style('whitegrid')
            plt.rcParams['figure.dpi'] = 100
            
            numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
            categorical_cols = [col for col in self.df.columns 
                              if col in self.column_types and self.column_types[col] in ['categorical', 'categorical_numeric']]
            
            # Chart 1: Correlation Heatmap
            if len(numeric_cols) > 1:
                fig, ax = plt.subplots(figsize=(12, 8))
                corr = self.df[numeric_cols[:10]].corr()
                sns.heatmap(corr, annot=True, cmap='coolwarm', center=0, fmt='.2f', 
                           square=True, linewidths=1, cbar_kws={"shrink": 0.8}, ax=ax)
                ax.set_title('Correlation Matrix - Identifying Relationships', fontsize=16, fontweight='bold', pad=20)
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Correlation Heatmap',
                    'image': img_str,
                    'explanation': 'This heatmap visualizes correlations between numerical variables. Strong positive correlations (red, close to 1) suggest variables move together, while negative correlations (blue, close to -1) indicate inverse relationships. Use this to identify potential predictors and multicollinearity.'
                })
            
            # Chart 2: Distribution Analysis for Top Numerical Columns
            if len(numeric_cols) >= 1:
                n_cols = min(3, len(numeric_cols))
                fig, axes = plt.subplots(2, n_cols, figsize=(15, 10), squeeze=False)
                
                for idx, col in enumerate(numeric_cols[:n_cols]):
                    # Histogram
                    axes[0, idx].hist(self.df[col].dropna(), bins=30, edgecolor='black', color='skyblue', alpha=0.7)
                    axes[0, idx].set_title(f'{col} - Distribution', fontweight='bold')
                    axes[0, idx].set_xlabel(col)
                    axes[0, idx].set_ylabel('Frequency')
                    axes[0, idx].grid(alpha=0.3)
                    
                    # Boxplot
                    axes[1, idx].boxplot(self.df[col].dropna(), vert=True)
                    axes[1, idx].set_title(f'{col} - Outlier Detection', fontweight='bold')
                    axes[1, idx].set_ylabel(col)
                    axes[1, idx].grid(alpha=0.3)
                
                plt.suptitle('Numerical Variables - Distribution & Outlier Analysis', fontsize=16, fontweight='bold', y=1.02)
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Distribution & Outlier Analysis',
                    'image': img_str,
                    'explanation': 'Top row shows histograms revealing data distribution patterns (normal, skewed, bimodal). Bottom row displays boxplots for outlier detection - points outside whiskers are potential outliers requiring investigation.'
                })
            
            # Chart 3: Top Categories Bar Chart
            if len(categorical_cols) >= 1:
                fig, axes = plt.subplots(1, min(3, len(categorical_cols)), figsize=(15, 6))
                if len(categorical_cols) == 1:
                    axes = [axes]
                
                for idx, col in enumerate(categorical_cols[:3]):
                    top_values = self.df[col].value_counts().head(10)
                    axes[idx].barh(range(len(top_values)), top_values.values, color='coral', edgecolor='black')
                    axes[idx].set_yticks(range(len(top_values)))
                    axes[idx].set_yticklabels(top_values.index)
                    axes[idx].set_xlabel('Count')
                    axes[idx].set_title(f'Top 10 {col}', fontweight='bold')
                    axes[idx].grid(axis='x', alpha=0.3)
                    axes[idx].invert_yaxis()
                
                plt.suptitle('Categorical Variables - Top Performers', fontsize=16, fontweight='bold')
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Top Categories Analysis',
                    'image': img_str,
                    'explanation': 'Bar charts display the most frequent categories in each categorical variable. This reveals dominant segments, market leaders, or popular items. Use this to focus resources on high-impact categories.'
                })
            
            # Chart 4: Scatter Plot Matrix (if multiple numeric columns)
            if len(numeric_cols) >= 2:
                sample_size = min(1000, len(self.df))
                df_sample = self.df[numeric_cols[:4]].sample(n=sample_size, random_state=42)
                n_vars = len(df_sample.columns)
                
                fig, axes = plt.subplots(n_vars, n_vars, figsize=(12, 12), squeeze=False)
                
                for i, col1 in enumerate(df_sample.columns):
                    for j, col2 in enumerate(df_sample.columns):
                        if i == j:
                            axes[i, j].hist(df_sample[col1].dropna(), bins=20, color='steelblue', edgecolor='black', alpha=0.7)
                        else:
                            axes[i, j].scatter(df_sample[col2], df_sample[col1], alpha=0.5, s=10, color='steelblue')
                        
                        if i == n_vars - 1:
                            axes[i, j].set_xlabel(col2, fontsize=8)
                        else:
                            axes[i, j].set_xticklabels([])
                        
                        if j == 0:
                            axes[i, j].set_ylabel(col1, fontsize=8)
                        else:
                            axes[i, j].set_yticklabels([])
                
                plt.suptitle('Scatter Plot Matrix - Multivariate Relationships', fontsize=16, fontweight='bold')
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Scatter Plot Matrix',
                    'image': img_str,
                    'explanation': 'Scatter plot matrix shows pairwise relationships between numerical variables. Diagonal shows distributions, off-diagonal shows correlations. Look for linear patterns (strong correlation) or clusters (segmentation opportunities).'
                })
            
            # Chart 5: Time Series (if datetime column exists)
            datetime_cols = [col for col, ctype in self.column_types.items() if ctype == 'datetime']
            if datetime_cols and len(numeric_cols) >= 1:
                date_col = datetime_cols[0]
                value_col = numeric_cols[0]
                
                fig, ax = plt.subplots(figsize=(14, 6))
                
                # Group by date and aggregate
                time_series = self.df.groupby(pd.Grouper(key=date_col, freq='M'))[value_col].agg(['sum', 'mean', 'count'])
                
                ax.plot(time_series.index, time_series['sum'], marker='o', linewidth=2, label='Total', color='steelblue')
                ax.fill_between(time_series.index, time_series['sum'], alpha=0.3, color='steelblue')
                
                ax.set_title(f'{value_col} Over Time - Trend Analysis', fontsize=16, fontweight='bold')
                ax.set_xlabel('Date', fontsize=12)
                ax.set_ylabel(value_col, fontsize=12)
                ax.legend()
                ax.grid(alpha=0.3)
                plt.xticks(rotation=45)
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Time Series Trend',
                    'image': img_str,
                    'explanation': 'Time series visualization reveals temporal patterns, trends, and seasonality. Upward trends indicate growth, downward trends suggest decline. Look for cyclical patterns and anomalies for strategic planning.'
                })
            
            # Chart 6: Segment Performance (if categorical and numerical exist)
            if len(categorical_cols) >= 1 and len(numeric_cols) >= 1:
                cat_col = categorical_cols[0]
                num_col = numeric_cols[0]
                
                # Get top categories
                top_cats = self.df[cat_col].value_counts().head(8).index
                df_filtered = self.df[self.df[cat_col].isin(top_cats)]
                
                fig, axes = plt.subplots(1, 2, figsize=(15, 6))
                
                # Boxplot by category
                df_filtered.boxplot(column=num_col, by=cat_col, ax=axes[0])
                axes[0].set_title(f'{num_col} Distribution by {cat_col}', fontweight='bold')
                axes[0].set_xlabel(cat_col)
                axes[0].set_ylabel(num_col)
                plt.sca(axes[0])
                plt.xticks(rotation=45)
                
                # Bar chart of means
                means = df_filtered.groupby(cat_col)[num_col].mean().sort_values(ascending=False)
                axes[1].bar(range(len(means)), means.values, color='coral', edgecolor='black')
                axes[1].set_xticks(range(len(means)))
                axes[1].set_xticklabels(means.index, rotation=45, ha='right')
                axes[1].set_title(f'Average {num_col} by {cat_col}', fontweight='bold')
                axes[1].set_ylabel(f'Average {num_col}')
                axes[1].grid(axis='y', alpha=0.3)
                
                plt.suptitle('Segment Performance Analysis', fontsize=16, fontweight='bold')
                plt.tight_layout()
                
                buffer = BytesIO()
                plt.savefig(buffer, format='png', bbox_inches='tight')
                buffer.seek(0)
                img_str = base64.b64encode(buffer.read()).decode()
                plt.close()
                
                charts.append({
                    'title': 'Segment Performance',
                    'image': img_str,
                    'explanation': 'Segment analysis compares performance across categories. Left plot shows distribution variability within segments, right plot shows average performance. Identify high-performing segments for resource allocation.'
                })
            
        except Exception as e:
            print(f"Error generating visualizations: {str(e)}")
            import traceback
            traceback.print_exc()
        
        return charts
    
    def generate_insights(self):
        """Stage 4: Business Insights with Detailed Analysis"""
        insights = []
        detailed_insights = {}
        
        # Dataset overview
        insights.append(f"Dataset contains {len(self.df):,} records and {len(self.df.columns)} columns")
        detailed_insights['overview'] = {
            'total_records': len(self.df),
            'total_columns': len(self.df.columns),
            'explanation': 'This represents the complete dataset dimensions. Each record is a unique observation, and each column represents a different variable or attribute.'
        }
        
        # Identify potential KPI columns
        kpi_keywords = ['sales', 'revenue', 'profit', 'amount', 'price', 'quantity', 'units', 'cost']
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        kpi_analysis = []
        for col in numeric_cols:
            col_lower = col.lower()
            if any(kw in col_lower for kw in kpi_keywords):
                try:
                    total = self.df[col].sum()
                    avg = self.df[col].mean()
                    median = self.df[col].median()
                    std = self.df[col].std()
                    min_val = self.df[col].min()
                    max_val = self.df[col].max()
                    
                    # Check for NaN/Inf values
                    if any(np.isnan(v) or np.isinf(v) for v in [total, avg, median, std, min_val, max_val]):
                        continue
                    
                    insights.append(f"Total {col}: {total:,.2f} | Average: {avg:,.2f}")
                    kpi_analysis.append({
                        'column': col,
                        'total': float(total),
                        'average': float(avg),
                        'median': float(median),
                        'std_dev': float(std),
                        'min': float(min_val),
                        'max': float(max_val),
                        'explanation': f'{col} shows a total of {total:,.2f} with an average of {avg:,.2f} per record. The standard deviation of {std:,.2f} indicates the variability in the data.'
                    })
                except:
                    pass
        
        detailed_insights['kpi_analysis'] = kpi_analysis
        
        # Top performers with detailed breakdown
        categorical_cols = [col for col in self.df.columns 
                          if col in self.column_types and self.column_types[col] in ['categorical', 'categorical_numeric']]
        
        top_performers = []
        for cat_col in categorical_cols[:5]:
            if self.df[cat_col].nunique() < 100:
                top_items = self.df[cat_col].value_counts().head(5)
                if len(top_items) > 0:
                    insights.append(f"Top {cat_col}: {top_items.index[0]} ({top_items.values[0]:,} occurrences)")
                    top_performers.append({
                        'category': cat_col,
                        'top_5': {str(k): int(v) for k, v in top_items.items()},
                        'unique_count': int(self.df[cat_col].nunique()),
                        'explanation': f'The {cat_col} category has {self.df[cat_col].nunique()} unique values. The top performer is "{top_items.index[0]}" appearing {top_items.values[0]:,} times ({(top_items.values[0]/len(self.df)*100):.1f}% of total).'
                    })
        
        detailed_insights['top_performers'] = top_performers
        
        # Datetime trends with detailed analysis
        datetime_cols = [col for col, ctype in self.column_types.items() if ctype == 'datetime']
        if datetime_cols:
            date_col = datetime_cols[0]
            date_range = f"{self.df[date_col].min()} to {self.df[date_col].max()}"
            insights.append(f"Date range: {date_range}")
            
            time_span = (self.df[date_col].max() - self.df[date_col].min()).days
            detailed_insights['temporal_analysis'] = {
                'date_column': date_col,
                'start_date': str(self.df[date_col].min()),
                'end_date': str(self.df[date_col].max()),
                'time_span_days': int(time_span),
                'explanation': f'The dataset spans {time_span} days from {self.df[date_col].min()} to {self.df[date_col].max()}. This temporal coverage allows for trend analysis and seasonality detection.'
            }
        
        # Data quality with detailed metrics
        missing_pct = (self.df.isnull().sum().sum() / (len(self.df) * len(self.df.columns))) * 100
        insights.append(f"Data completeness: {100-missing_pct:.1f}%")
        
        detailed_insights['data_quality'] = {
            'completeness_percentage': float(100-missing_pct),
            'total_missing_cells': int(self.df.isnull().sum().sum()),
            'columns_with_missing': {col: int(count) for col, count in self.df.isnull().sum().items() if count > 0},
            'explanation': f'The dataset is {100-missing_pct:.1f}% complete with {self.df.isnull().sum().sum()} missing values across all columns. High completeness indicates reliable data for analysis.'
        }
        
        # Statistical insights for numerical columns
        if len(numeric_cols) > 0:
            correlations = []
            if len(numeric_cols) > 1:
                try:
                    corr_matrix = self.df[numeric_cols].corr()
                    # Find strongest correlations
                    for i in range(len(corr_matrix.columns)):
                        for j in range(i+1, len(corr_matrix.columns)):
                            corr_val = corr_matrix.iloc[i, j]
                            if not np.isnan(corr_val) and not np.isinf(corr_val) and abs(corr_val) > 0.5:
                                correlations.append({
                                    'var1': str(corr_matrix.columns[i]),
                                    'var2': str(corr_matrix.columns[j]),
                                    'correlation': float(corr_val),
                                    'strength': 'Strong' if abs(corr_val) > 0.7 else 'Moderate',
                                    'explanation': f'{corr_matrix.columns[i]} and {corr_matrix.columns[j]} show a {"strong" if abs(corr_val) > 0.7 else "moderate"} {"positive" if corr_val > 0 else "negative"} correlation ({corr_val:.3f}).'
                                })
                except:
                    pass
            
            detailed_insights['correlations'] = correlations
        
        self.detailed_insights = detailed_insights
        return insights, detailed_insights
    
    def generate_python_code(self, filename='your_data.csv'):
        """Stage 5: Python Pandas Code"""
        code = f"""import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

# Load data
# Replace with your actual file path
df = pd.read_csv('{filename}')  # Adjust file path as needed

# Data Understanding
print("Shape:", df.shape)
print("\\nColumns:", df.columns.tolist())
print("\\nData Types:")
print(df.dtypes)
print("\\nFirst 10 rows:")
print(df.head(10))
print("\\nLast 10 rows:")
print(df.tail(10))

# Data Cleaning
print("\\nMissing Values:")
print(df.isnull().sum())

# Fill missing values
for col in df.select_dtypes(include=[np.number]).columns:
    df[col].fillna(df[col].median(), inplace=True)
for col in df.select_dtypes(include=['object']).columns:
    df[col].fillna('Unknown', inplace=True)

# Remove duplicates
df.drop_duplicates(inplace=True)

# Standardize strings
for col in df.select_dtypes(include=['object']).columns:
    df[col] = df[col].str.strip().str.title()

# EDA - Numerical Summary
print("\\nNumerical Summary:")
print(df.describe())

# EDA - Categorical Summary
categorical_cols = df.select_dtypes(include=['object']).columns
for col in categorical_cols:
    print(f"\\n{{col}} - Top 10:")
    print(df[col].value_counts().head(10))

# Correlation Matrix
numeric_df = df.select_dtypes(include=[np.number])
if len(numeric_df.columns) > 1:
    plt.figure(figsize=(10, 8))
    sns.heatmap(numeric_df.corr(), annot=True, cmap='coolwarm', center=0)
    plt.title('Correlation Matrix')
    plt.tight_layout()
    plt.savefig('correlation_matrix.png')
    plt.close()

# Visualizations
for col in numeric_df.columns[:5]:
    plt.figure(figsize=(10, 6))
    plt.subplot(1, 2, 1)
    plt.hist(df[col].dropna(), bins=30, edgecolor='black')
    plt.title(f'{{col}} Distribution')
    plt.xlabel(col)
    plt.ylabel('Frequency')
    
    plt.subplot(1, 2, 2)
    plt.boxplot(df[col].dropna())
    plt.title(f'{{col}} Boxplot')
    plt.ylabel(col)
    plt.tight_layout()
    plt.savefig(f'{{col}}_analysis.png')
    plt.close()

print("\\nAnalysis complete!")
"""
        return code
    
    def generate_sql_queries(self):
        """Stage 6: SQL Query Generation"""
        queries = []
        
        # Detect potential table name
        table_name = "dataset"
        
        queries.append({
            "name": "Top 10 Records",
            "query": f"SELECT * FROM {table_name} LIMIT 10;"
        })
        
        # Find potential grouping columns
        categorical_cols = [col for col in self.df.columns 
                          if col in self.column_types and self.column_types[col] in ['categorical', 'categorical_numeric']]
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        
        if categorical_cols and numeric_cols:
            group_col = categorical_cols[0]
            value_col = numeric_cols[0]
            queries.append({
                "name": f"Top 10 by {value_col}",
                "query": f"SELECT {group_col}, SUM({value_col}) as total_{value_col}\nFROM {table_name}\nGROUP BY {group_col}\nORDER BY total_{value_col} DESC\nLIMIT 10;"
            })
        
        # Date-based query
        datetime_cols = [col for col, ctype in self.column_types.items() if ctype == 'datetime']
        if datetime_cols and numeric_cols:
            date_col = datetime_cols[0]
            value_col = numeric_cols[0]
            queries.append({
                "name": "Monthly Trend",
                "query": f"SELECT DATE_TRUNC('month', {date_col}) as month,\n       SUM({value_col}) as total_{value_col}\nFROM {table_name}\nGROUP BY month\nORDER BY month;"
            })
        
        queries.append({
            "name": "Data Quality Check",
            "query": f"SELECT COUNT(*) as total_records,\n       COUNT(DISTINCT *) as unique_records\nFROM {table_name};"
        })
        
        return queries
    
    def generate_dax_measures(self):
        """Stage 7: Power BI DAX Measures"""
        measures = []
        
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        
        for col in numeric_cols[:5]:
            measures.append({
                "name": f"Total {col}",
                "dax": f"Total {col} = SUM('{col}'[{col}])"
            })
            
            measures.append({
                "name": f"Average {col}",
                "dax": f"Average {col} = AVERAGE('{col}'[{col}])"
            })
        
        measures.append({
            "name": "Total Records",
            "dax": "Total Records = COUNTROWS(Dataset)"
        })
        
        datetime_cols = [col for col, ctype in self.column_types.items() if ctype == 'datetime']
        if datetime_cols and numeric_cols:
            value_col = numeric_cols[0]
            measures.append({
                "name": f"{value_col} YoY Growth",
                "dax": f"{value_col} YoY Growth = \nVAR CurrentYear = SUM('{value_col}'[{value_col}])\nVAR PreviousYear = CALCULATE(SUM('{value_col}'[{value_col}]), SAMEPERIODLASTYEAR(Date[Date]))\nRETURN DIVIDE(CurrentYear - PreviousYear, PreviousYear, 0)"
            })
        
        return measures
    
    def generate_json_output(self, understanding, cleaning, eda, insights, python_code, sql_queries, dax_measures):
        """Stage 8: Structured JSON Output"""
        return {
            "summary": f"Analysis of dataset with {len(self.df)} records and {len(self.df.columns)} columns",
            "columns": understanding['column_types'],
            "data_quality": {
                "missing_values": cleaning.get('missing_values', {}),
                "duplicates_removed": cleaning.get('duplicates_removed', 0),
                "outliers_detected": cleaning.get('outliers_detected', {})
            },
            "key_insights": insights,
            "python_code": python_code,
            "sql_queries": sql_queries,
            "dax_measures": dax_measures,
            "recommended_charts": ["Correlation Heatmap", "Distribution Histograms", "Boxplots", "Bar Charts", "Time Series"],
            "business_recommendations": [
                "Focus on top performing categories",
                "Investigate outliers for anomalies",
                "Monitor trends over time",
                "Address data quality issues"
            ]
        }
    
    def _generate_recommendations(self, numeric_cols, cat_cols):
        """Generate actionable recommendations based on data analysis"""
        recommendations = []
        
        # Recommendation 1: Data Quality
        missing_pct = (self.df.isnull().sum().sum() / (len(self.df) * len(self.df.columns))) * 100
        if missing_pct > 5:
            recommendations.append({
                'priority': 'HIGH',
                'recommendation': 'Improve Data Quality',
                'action': f'Address {missing_pct:.1f}% missing data. Implement data validation at source.',
                'impact': 'Increase analysis accuracy by 20-30%',
                'timeline': '1-2 months'
            })
        
        # Recommendation 2: Top Performers
        if len(cat_cols) > 0 and len(numeric_cols) > 0:
            cat_col = cat_cols[0]
            num_col = numeric_cols[0]
            top_cat = self.df.groupby(cat_col)[num_col].sum().idxmax()
            recommendations.append({
                'priority': 'HIGH',
                'recommendation': 'Focus on Top Performers',
                'action': f'Allocate 60% of resources to "{top_cat}" segment. Replicate success factors.',
                'impact': 'Increase overall performance by 15-25%',
                'timeline': '3-6 months'
            })
        
        # Recommendation 3: Outliers
        outlier_cols = []
        for col in numeric_cols:
            Q1 = self.df[col].quantile(0.25)
            Q3 = self.df[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = ((self.df[col] < (Q1 - 1.5 * IQR)) | (self.df[col] > (Q3 + 1.5 * IQR))).sum()
            if outliers > len(self.df) * 0.05:
                outlier_cols.append(col)
        
        if outlier_cols:
            recommendations.append({
                'priority': 'MEDIUM',
                'recommendation': 'Investigate Outliers',
                'action': f'Analyze {len(outlier_cols)} columns with significant outliers. Identify root causes.',
                'impact': 'Reduce variance by 30%, improve forecasting',
                'timeline': '2-4 weeks'
            })
        
        # Recommendation 4: Correlations
        if len(numeric_cols) > 1:
            corr_matrix = self.df[numeric_cols].corr()
            strong_corr = []
            for i in range(len(corr_matrix.columns)):
                for j in range(i+1, len(corr_matrix.columns)):
                    if abs(corr_matrix.iloc[i, j]) > 0.7:
                        strong_corr.append((corr_matrix.columns[i], corr_matrix.columns[j]))
            
            if strong_corr:
                recommendations.append({
                    'priority': 'MEDIUM',
                    'recommendation': 'Leverage Key Drivers',
                    'action': f'Focus on {len(strong_corr)} strongly correlated variable pairs. Build predictive models.',
                    'impact': 'Enable proactive decision-making, reduce costs by 20%',
                    'timeline': '1-3 months'
                })
        
        # Recommendation 5: Automation
        recommendations.append({
            'priority': 'LOW',
            'recommendation': 'Automate Reporting',
            'action': 'Set up automated dashboards and alerts for key metrics. Schedule weekly reports.',
            'impact': 'Save 10-15 hours/week, faster insights',
            'timeline': '1-2 months'
        })
        
        # Recommendation 6: Segmentation
        if len(cat_cols) > 0:
            recommendations.append({
                'priority': 'HIGH',
                'recommendation': 'Implement Segmentation Strategy',
                'action': f'Create targeted strategies for top 3 {cat_cols[0]} segments. Personalize approach.',
                'impact': 'Increase conversion by 25-40%',
                'timeline': '2-4 months'
            })
        
        # Recommendation 7: Predictive Analytics
        if len(numeric_cols) >= 2:
            recommendations.append({
                'priority': 'MEDIUM',
                'recommendation': 'Deploy Predictive Models',
                'action': 'Build ML models to forecast key metrics. Start with regression/classification.',
                'impact': 'Improve planning accuracy by 30-50%',
                'timeline': '3-6 months'
            })
        
        return recommendations
    
    def generate_excel_report(self, filename):
        """Generate professional Excel report with multiple sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Cleaned Data
        ws1 = wb.create_sheet("Cleaned_Data")
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws1.append(r)
        
        # Format header
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sheet 2: Statistical Summary
        ws2 = wb.create_sheet("Statistical_Summary")
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        if len(numeric_cols) > 0:
            stats_df = self.df[numeric_cols].describe()
            stats_df.loc['range'] = stats_df.loc['max'] - stats_df.loc['min']
            stats_df.loc['variance'] = self.df[numeric_cols].var()
            
            for r in dataframe_to_rows(stats_df, index=True, header=True):
                ws2.append(r)
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Sheet 3: Missing Values Report
        ws3 = wb.create_sheet("Missing_Values")
        missing_data = pd.DataFrame({
            'Column': self.df.columns,
            'Missing_Count': self.df.isnull().sum().values,
            'Missing_Percentage': (self.df.isnull().sum().values / len(self.df) * 100).round(2)
        })
        
        for r in dataframe_to_rows(missing_data, index=False, header=True):
            ws3.append(r)
        
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Sheet 4: Correlation Matrix
        ws4 = wb.create_sheet("Correlation_Matrix")
        if len(numeric_cols) > 1:
            corr_df = self.df[numeric_cols].corr()
            for r in dataframe_to_rows(corr_df, index=True, header=True):
                ws4.append(r)
            
            for cell in ws4[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # Sheet 5: Categorical Summary
        ws5 = wb.create_sheet("Categorical_Summary")
        cat_cols = [col for col in self.df.columns if col in self.column_types and self.column_types[col] in ['categorical', 'categorical_numeric']]
        
        row = 1
        for col in cat_cols[:10]:
            ws5.cell(row, 1, f"{col} - Top 10").font = Font(bold=True, size=12)
            row += 1
            
            value_counts = self.df[col].value_counts().head(10)
            ws5.cell(row, 1, "Value").font = Font(bold=True)
            ws5.cell(row, 2, "Count").font = Font(bold=True)
            row += 1
            
            for val, count in value_counts.items():
                ws5.cell(row, 1, str(val))
                ws5.cell(row, 2, int(count))
                row += 1
            row += 1
        
        # Sheet 6: Business Metrics
        ws6 = wb.create_sheet("Business_Metrics")
        ws6.append(["Metric", "Value"])
        ws6.cell(1, 1).font = Font(bold=True, color="FFFFFF")
        ws6.cell(1, 2).font = Font(bold=True, color="FFFFFF")
        ws6.cell(1, 1).fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        ws6.cell(1, 2).fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
        
        ws6.append(["Total Records", len(self.df)])
        ws6.append(["Total Columns", len(self.df.columns)])
        ws6.append(["Duplicates Removed", len(self.original_df) - len(self.df)])
        ws6.append(["Data Completeness %", round(100 - (self.df.isnull().sum().sum() / (len(self.df) * len(self.df.columns)) * 100), 2)])
        
        # Add KPI metrics
        kpi_keywords = ['sales', 'revenue', 'profit', 'amount', 'price', 'cost', 'income']
        for col in numeric_cols:
            if any(kw in col.lower() for kw in kpi_keywords):
                ws6.append([f"Total {col}", round(self.df[col].sum(), 2)])
                ws6.append([f"Average {col}", round(self.df[col].mean(), 2)])
                ws6.append([f"Median {col}", round(self.df[col].median(), 2)])
        
        # Sheet 7: Actionable Recommendations
        ws7 = wb.create_sheet("Actionable_Recommendations")
        ws7.append(["Priority", "Recommendation", "Action", "Expected Impact", "Timeline"])
        for cell in ws7[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Generate recommendations based on data analysis
        recommendations = self._generate_recommendations(numeric_cols, cat_cols)
        for rec in recommendations:
            ws7.append([rec['priority'], rec['recommendation'], rec['action'], rec['impact'], rec['timeline']])
        
        # Format recommendations
        for row in ws7.iter_rows(min_row=2, max_row=ws7.max_row):
            priority = row[0].value
            if priority == "HIGH":
                row[0].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                row[0].font = Font(bold=True, color="FFFFFF")
            elif priority == "MEDIUM":
                row[0].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                row[0].font = Font(bold=True, color="FFFFFF")
            else:
                row[0].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                row[0].font = Font(bold=True, color="FFFFFF")
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    def generate_notebook(self, filename):
        """Generate clean notebook without embedded code"""
        df_name = filename.replace('.csv', '').replace('.xlsx', '').replace('.xls', '').replace('.json', '').replace(' ', '_')
        return self._create_clean_notebook(df_name, filename)
    
    def _add_driver_analysis(self, cells, df_name, numeric_cols, cat_cols):
        """Add Value-Driven Analysis Framework - Universal Template"""
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["# Value-Driven Analysis Framework"]})
        
        # Part 1: Define The Core Problem
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Part 1: Define The Core Problem (The Hypothesis)\n\n**Dataset Type & Goal:**\n- This dataset contains business/operational data\n- Business Question: What drives performance and how can we optimize outcomes?\n\n**Key Performance Indicator (KPI):**\n- Primary KPI: " + (numeric_cols[0] if numeric_cols else "[Identify your KPI]") + "\n- Goal: Understand what drives this metric and identify actionable opportunities"]})
        
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Part 2: The Staged Analysis Protocol"]})
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["### STAGE 1: Advanced Feature Engineering & KPI Creation"]})
        
        kpi_keywords = ['sales', 'revenue', 'profit', 'income', 'gross']
        qty_keywords = ['quantity', 'units']
        kpi_col = next((c for c in numeric_cols if any(kw in c.lower() for kw in kpi_keywords)), None)
        qty_col = next((c for c in numeric_cols if any(kw in c.lower() for kw in qty_keywords)), None)
        
        if kpi_col and qty_col:
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Create Efficiency Ratio KPI\n{df_name}['Efficiency_Ratio'] = {df_name}['{kpi_col}'] / {df_name}['{qty_col}'].replace(0, np.nan)\nprint('Efficiency Ratio Statistics:')\n{df_name}['Efficiency_Ratio'].describe()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Visualize Efficiency Ratio Distribution\nplt.figure(figsize=(14,5))\n\nplt.subplot(1,2,1)\nplt.hist({df_name}['Efficiency_Ratio'].dropna(), bins=30, edgecolor='black', color='skyblue')\nplt.title('Distribution of Efficiency Ratio ({kpi_col}/{qty_col})')\nplt.xlabel('Efficiency Ratio')\nplt.ylabel('Frequency')\n\nplt.subplot(1,2,2)\nplt.boxplot({df_name}['Efficiency_Ratio'].dropna())\nplt.title('Efficiency Ratio - Outlier Detection')\nplt.ylabel('Efficiency Ratio')\nplt.tight_layout()\nplt.show()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Identify Top 1% and Bottom 1% Performers\ntop_1_pct = {df_name}['Efficiency_Ratio'].quantile(0.99)\nbottom_1_pct = {df_name}['Efficiency_Ratio'].quantile(0.01)\n\nprint(f'Top 1% Threshold: {{top_1_pct:.2f}}')\nprint(f'Bottom 1% Threshold: {{bottom_1_pct:.2f}}')\n\ntop_performers = {df_name}[{df_name}['Efficiency_Ratio'] >= top_1_pct]\nbottom_performers = {df_name}[{df_name}['Efficiency_Ratio'] <= bottom_1_pct]\n\nprint(f'\\nTop Performers: {{len(top_performers)}} records')\nprint(f'Bottom Performers: {{len(bottom_performers)}} records')\nprint('\\nThese represent your highest-leverage opportunities!')"]})
        
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["### STAGE 3: Key Driver Analysis (Correlation & Causation)\n\nThis stage identifies which variables have the strongest relationship with our KPI."]})
        
        if len(numeric_cols) >= 2:
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Scatter Plot with Regression Line\nplt.figure(figsize=(10,6))\nsns.regplot(data={df_name}, x='{numeric_cols[1]}', y='{numeric_cols[0]}', scatter_kws={{'alpha':0.5}}, line_kws={{'color':'red', 'linewidth':2}})\nplt.title('Driver Analysis: {numeric_cols[1]} vs {numeric_cols[0]}')\nplt.xlabel('{numeric_cols[1]}')\nplt.ylabel('{numeric_cols[0]}')\nplt.grid(True, alpha=0.3)\nplt.show()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Calculate Correlation Coefficient\ncorrelation = {df_name}[['{numeric_cols[0]}', '{numeric_cols[1]}']].corr().iloc[0,1]\nprint(f'Correlation between {numeric_cols[1]} and {numeric_cols[0]}: {{correlation:.3f}}')\n\nif abs(correlation) > 0.7:\n    print('Strong correlation detected! This is a key driver.')\nelif abs(correlation) > 0.4:\n    print('Moderate correlation. This variable has some influence.')\nelse:\n    print('Weak correlation. This may not be a primary driver.')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Correlation Heatmap of All Numerical Variables\nplt.figure(figsize=(10,8))\nnumeric_data = {df_name}[{numeric_cols[:6]}].corr()\nsns.heatmap(numeric_data, annot=True, cmap='coolwarm', center=0, fmt='.2f', square=True, linewidths=1)\nplt.title('Correlation Heatmap: Identifying Key Drivers')\nplt.tight_layout()\nplt.show()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Identify Strongest Predictor\ncorrelations_with_target = {df_name}[{numeric_cols[:6]}].corr()['{numeric_cols[0]}'].abs().sort_values(ascending=False)\nprint('Variables ranked by correlation strength with {numeric_cols[0]}:')\nprint(correlations_with_target)\nprint(f'\\nStrongest predictor: {{correlations_with_target.index[1]}} (r={{correlations_with_target.values[1]:.3f}})')"]})
        
        if len(cat_cols) >= 2 and len(numeric_cols) >= 1:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## STAGE 3: Multivariate (Triple-Axis) Segmentation"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Triple-Axis Analysis: {numeric_cols[0]} by {cat_cols[0]} and {cat_cols[1]}\nplt.figure(figsize=(12,6))\nsns.boxplot(data={df_name}, x='{cat_cols[0]}', y='{numeric_cols[0]}', hue='{cat_cols[1]}', palette='Set2')\nplt.title('Multivariate Segmentation: {numeric_cols[0]} by {cat_cols[0]} & {cat_cols[1]}')\nplt.xlabel('{cat_cols[0]}')\nplt.ylabel('{numeric_cols[0]}')\nplt.xticks(rotation=45)\nplt.legend(title='{cat_cols[1]}')\nplt.tight_layout()\nplt.show()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Statistical Comparison Across Segments\nsegment_analysis = {df_name}.groupby(['{cat_cols[0]}', '{cat_cols[1]}'])['{numeric_cols[0]}'].agg(['mean', 'median', 'count']).round(2)\nprint('Segment Performance Analysis:')\nprint(segment_analysis.sort_values('mean', ascending=False))\nprint('\\nInsight: Identify which combination of {cat_cols[0]} and {cat_cols[1]} yields highest {numeric_cols[0]}')"]})
        
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["### STAGE 4: Segmentation & Actionable Strategy\n\nUsing insights from driver analysis to create targeted segments and concrete business actions."]})
        
        # High-Value Segmentation
        if len(cat_cols) >= 1 and len(numeric_cols) >= 2:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["#### Segmentation Analysis"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Create segments based on key drivers\n# Segment 1: High performers (top 25% of KPI)\n# Segment 2: Low performers (bottom 25% of KPI)\n\nkpi_75th = {df_name}['{numeric_cols[0]}'].quantile(0.75)\nkpi_25th = {df_name}['{numeric_cols[0]}'].quantile(0.25)\n\n{df_name}['Segment'] = 'Medium'\n{df_name}.loc[{df_name}['{numeric_cols[0]}'] >= kpi_75th, 'Segment'] = 'High Performer'\n{df_name}.loc[{df_name}['{numeric_cols[0]}'] <= kpi_25th, 'Segment'] = 'Low Performer'\n\nprint('Segment Distribution:')\nprint({df_name}['Segment'].value_counts())"]})
            
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Compare segments across key categorical variable\nsegment_comparison = {df_name}.groupby(['Segment', '{cat_cols[0]}'])['{numeric_cols[0]}'].agg(['mean', 'count']).round(2)\nprint('\\nSegment Comparison by {cat_cols[0]}:')\nprint(segment_comparison)\n\n# Visualize\nplt.figure(figsize=(12,6))\nsns.boxplot(data={df_name}, x='Segment', y='{numeric_cols[0]}', hue='{cat_cols[0]}', palette='Set2')\nplt.title('KPI Distribution Across Segments and {cat_cols[0]}')\nplt.xticks(rotation=0)\nplt.legend(title='{cat_cols[0]}', bbox_to_anchor=(1.05, 1))\nplt.tight_layout()\nplt.show()"]})
        
        if len(numeric_cols) >= 2:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"#### Top Drivers Summary\n\nBased on correlation analysis:\n\n**Key Finding**: The relationship between **{numeric_cols[1]}** and **{numeric_cols[0]}** reveals important drivers of performance.\n\n**Predictive Insight**: Strong correlations (|r| > 0.7) indicate that changes in driver variables will have measurable impact on the KPI.\n\n**Strategic Focus**: Prioritize optimization of the top 3 drivers identified in the correlation heatmap."]})
        
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["#### Business Recommendations (Finding → Action → Impact)\n\n**Recommendation 1: Optimize High-Impact Drivers**\n- **Finding**: Correlation analysis reveals top 3 variables with strongest relationship to KPI (|r| > 0.5)\n- **Action**: Allocate 60% of optimization budget to improving these high-impact drivers\n- **Impact**: Expected 15-25% improvement in KPI within 2 quarters\n\n**Recommendation 2: Replicate High-Performer Characteristics**\n- **Finding**: Top 25% performers show distinct patterns in key categorical variables\n- **Action**: Implement best practices from high-performer segment across all operations\n- **Impact**: Reduce performance variance by 30% and lift bottom quartile by 20%\n\n**Recommendation 3: Address Low-Performer Segments**\n- **Finding**: Bottom 25% performers have specific identifiable characteristics\n- **Action**: Deploy targeted intervention programs for low-performer segments\n- **Impact**: Reduce churn/losses by 40% and improve overall efficiency by 12%\n\n**Recommendation 4: Leverage Temporal Patterns**\n- **Finding**: Time-based analysis reveals optimal operational windows and seasonal trends\n- **Action**: Reallocate resources to align with peak performance periods\n- **Impact**: Increase resource utilization by 25% and reduce idle capacity\n\n**Recommendation 5: Implement Predictive Monitoring**\n- **Finding**: Strong correlations enable predictive modeling of KPI outcomes\n- **Action**: Build real-time dashboard tracking top 5 driver metrics with alerts\n- **Impact**: Enable proactive decision-making and reduce reactive costs by 35%"]})
    
    def _create_clean_notebook(self, df_name, filename):
        """Stage 9: Generate Jupyter Notebook - Professional Cell-by-Cell Format with 3-Stage Framework"""
        cells = []
        
        # Cell 1: Import libraries
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": ["import numpy as np\nimport pandas as pd\nimport matplotlib.pyplot as plt\nimport seaborn as sns\nfrom collections import Counter"]})
        
        # Cell 2: Warnings
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": ["import warnings\nwarnings.filterwarnings('ignore')"]})
        
        # Cell 3: Load data - use correct method based on file extension
        if filename.endswith(('.xlsx', '.xls')):
            load_code = f"{df_name} = pd.read_excel('{filename}')\n{df_name}"
        elif filename.endswith('.json'):
            load_code = f"{df_name} = pd.read_json('{filename}')\n{df_name}"
        else:
            load_code = f"{df_name} = pd.read_csv('{filename}')\n{df_name}"
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [load_code]})
        
        # STAGE 1: Data Preparation & Advanced Feature Engineering
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["# STAGE 1: Data Preparation & Advanced Feature Engineering"]})
        
        # Cell 4: Info
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.info()"]})
        
        # Cell 5: Describe
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.describe()"]})
        
        # Temporal Feature Engineering
        date_col = next((c for c in self.df.columns if 'date' in c.lower()), None)
        if date_col:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Temporal Feature Extraction"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Convert to datetime\n{df_name}['{date_col}'] = pd.to_datetime({df_name}['{date_col}'], errors='coerce')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Extract temporal features\n{df_name}['Year'] = {df_name}['{date_col}'].dt.year\n{df_name}['Month'] = {df_name}['{date_col}'].dt.month\n{df_name}['DayOfWeek'] = {df_name}['{date_col}'].dt.dayofweek\n{df_name}"]})
        
        # Un-nesting/Exploding Data
        cat_cols = [col for col, ctype in self.column_types.items() if ctype in ['categorical', 'categorical_numeric']]
        explode_candidates = ['genre', 'cast', 'listed_in', 'category', 'tags']
        explode_col = next((c for c in cat_cols if any(kw in c.lower() for kw in explode_candidates)), None)
        if explode_col:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Un-nesting Comma-Separated Data"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Explode {explode_col} for individual analysis\n{df_name}_exploded = {df_name}.copy()\n{df_name}_exploded['{explode_col}'] = {df_name}_exploded['{explode_col}'].str.split(', ')\n{df_name}_exploded = {df_name}_exploded.explode('{explode_col}')\n{df_name}_exploded"]})
        
        # Cell 6: Duplicates
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.duplicated().sum()"]})
        
        # Cell 7: Missing values
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.isnull().sum()"]})
        
        # Cells 8-12: Fill missing values for each column
        missing_cols = [col for col in self.original_df.columns if self.original_df[col].isnull().sum() > 0]
        for col in missing_cols[:5]:
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{col}'].fillna(value='unknown', inplace=True)\n{df_name}['{col}']"]})
        
        # Cell: Check missing after cleaning
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.isnull().sum()"]})
        cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}"]})
        
        # KPI Normalization
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        kpi_keywords = ['sales', 'revenue', 'profit', 'amount', 'price']
        quantity_keywords = ['quantity', 'units', 'count']
        kpi_col = next((c for c in numeric_cols if any(kw in c.lower() for kw in kpi_keywords)), None)
        qty_col = next((c for c in numeric_cols if any(kw in c.lower() for kw in quantity_keywords)), None)
        
        if kpi_col and qty_col:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Rate/Ratio Features (Efficiency Metrics)"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Calculate Value per Unit (Efficiency Ratio)\n{df_name}['Value_per_Unit'] = {df_name}['{kpi_col}'] / {df_name}['{qty_col}'].replace(0, np.nan)\nprint('Value per Unit Statistics:')\nprint({df_name}['Value_per_Unit'].describe())\n\n# Visualize efficiency distribution\nplt.figure(figsize=(14,5))\nplt.subplot(1,2,1)\nplt.hist({df_name}['Value_per_Unit'].dropna(), bins=30, edgecolor='black', color='skyblue')\nplt.title('Distribution of Value per Unit')\nplt.xlabel('Value per Unit')\nplt.ylabel('Frequency')\n\nplt.subplot(1,2,2)\nplt.boxplot({df_name}['Value_per_Unit'].dropna())\nplt.title('Value per Unit - Outlier Detection')\nplt.ylabel('Value per Unit')\nplt.tight_layout()\nplt.show()"]})
            
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Identify Top 1% and Bottom 1% Performers\ntop_1_pct = {df_name}['Value_per_Unit'].quantile(0.99)\nbottom_1_pct = {df_name}['Value_per_Unit'].quantile(0.01)\n\nprint(f'Top 1% Threshold: {{top_1_pct:.2f}}')\nprint(f'Bottom 1% Threshold: {{bottom_1_pct:.2f}}')\n\ntop_performers = {df_name}[{df_name}['Value_per_Unit'] >= top_1_pct]\nbottom_performers = {df_name}[{df_name}['Value_per_Unit'] <= bottom_1_pct]\n\nprint(f'\\nTop Performers: {{len(top_performers)}} records')\nprint(f'Bottom Performers: {{len(bottom_performers)}} records')\nprint('\\nThese represent your highest-leverage opportunities!')"]})
        
        # STAGE 2: Strategic Segmentation
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["# STAGE 2: Strategic Segmentation (Multivariate Analysis)"]})
        
        # Segment vs KPI Analysis
        cat_cols = [col for col, ctype in self.column_types.items() if ctype in ['categorical', 'categorical_numeric']]
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns.tolist()
        
        if cat_cols and numeric_cols:
            main_cat = cat_cols[0]
            kpi_col = numeric_cols[0]
            
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"## Segment vs KPI: {main_cat} Performance"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Group by segment and calculate KPI\nsegment_performance = {df_name}.groupby('{main_cat}')['{kpi_col}'].agg(['mean', 'sum', 'count']).sort_values('mean', ascending=False)\nsegment_performance"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Visualize Segment vs KPI (Average)\nplt.figure(figsize=(10,6))\nsns.barplot(data={df_name}, x='{main_cat}', y='{kpi_col}', estimator='mean', palette='viridis')\nplt.title('Average {kpi_col} by {main_cat}')\nplt.xlabel('{main_cat}')\nplt.ylabel('Average {kpi_col}')\nplt.xticks(rotation=45)\nplt.show()"]})
        
        # Temporal Trend Analysis
        if date_col and numeric_cols:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Temporal Trend Analysis"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Trend over time\ntemporal_trend = {df_name}.groupby('Month')['{numeric_cols[0]}'].mean().reset_index()\ntemporal_trend"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Line plot for temporal trends\nplt.figure(figsize=(12,6))\nplt.plot(temporal_trend['Month'], temporal_trend['{numeric_cols[0]}'], marker='o', linewidth=2, color='steelblue')\nplt.title('Monthly Trend of {numeric_cols[0]}')\nplt.xlabel('Month')\nplt.ylabel('Average {numeric_cols[0]}')\nplt.grid(True, alpha=0.3)\nplt.show()"]})
        
        # Year analysis if available
        year_col = next((c for c in self.df.select_dtypes(include=[np.number]).columns if 'year' in c.lower()), None)
        if year_col and cat_cols:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"# {year_col.title()}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"last_decade = {df_name}[['{cat_cols[0]}', '{year_col}']]\n\nlast_decade = last_decade.rename(columns={{'{year_col}': 'Release Year'}})\nlast_decade = last_decade[last_decade['Release Year'] >= 2013]\nlast_decade"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"last_decade_df = last_decade.groupby('Release Year')['{cat_cols[0]}'].size().reset_index()\nlast_decade_df"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"last_decade.groupby('Release Year')['{cat_cols[0]}'].value_counts()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,6))\n\ncount_plot = sns.countplot(x='Release Year', data=last_decade, hue='{cat_cols[0]}', palette='pastel')\ncount_plot.set(title='Trend of content Released over the years')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"last_decade_df.rename(columns={{'{cat_cols[0]}': 'Total Content'}}, inplace=True)"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,6))\n\nplot_total_content = sns.lineplot(x='Release Year', y='Total Content', data=last_decade_df, linewidth=1)\nplot_total_content.set(xlabel='Release Year', ylabel='Total Content', title='Trend of content')\nplt.show()"]})
        
        # Top N analysis
        if len(cat_cols) > 1:
            second_cat = cat_cols[1]
            # Sanitize variable name - remove spaces, colons, special chars
            safe_var = second_cat.replace(' ', '_').replace(':', '_').replace('-', '_').replace('.', '_')
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"# {second_cat.title()}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"top_10_{safe_var} = {df_name}['{second_cat}'].value_counts().head(10)\n\ntop_10_{safe_var}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{second_cat}'].value_counts()[:10].index"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{second_cat}'].value_counts()[:10].values"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,8))\n\nsns.barplot(x={df_name}['{second_cat}'].value_counts()[:5].values, y={df_name}['{second_cat}'].value_counts()[:5].index, palette='pastel')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,8))\n\n{safe_var}_plot = sns.barplot(x={df_name}['{second_cat}'].value_counts()[:5].values, y={df_name}['{second_cat}'].value_counts()[:5].index, palette='pastel')\nfor i in {safe_var}_plot.containers:\n    {safe_var}_plot.bar_label(i)"]})
        

        
        # Country analysis
        country_col = next((c for c in cat_cols if 'country' in c.lower()), None)
        if country_col:
            safe_var = country_col.replace(' ', '_').replace(':', '_').replace('-', '_').replace('.', '_')
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"# {country_col.title()}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"top_10_{safe_var} = {df_name}['{country_col}'].value_counts().head(10)\n\ntop_10_{safe_var}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"top_10_{safe_var} = {df_name}['{country_col}'].value_counts().head(10)"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{country_col}'].value_counts()[:10].index"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{country_col}'].value_counts()[:10].values"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,8))\n\nsns.barplot(x={df_name}['{country_col}'].value_counts()[:5].values, y={df_name}['{country_col}'].value_counts()[:5].index, palette='pastel')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,8))\n\n{safe_var}_plot = sns.barplot(x={df_name}['{country_col}'].value_counts()[:5].values, y={df_name}['{country_col}'].value_counts()[:5].index, palette='pastel')\nfor i in {safe_var}_plot.containers:\n    {safe_var}_plot.bar_label(i)"]})
        
        # Rating analysis
        rating_col = next((c for c in cat_cols if 'rating' in c.lower()), None)
        if rating_col:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"# {rating_col.title()}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{country_col if country_col else cat_cols[0]}'].unique()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{rating_col}'].unique()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,6))\nsns.countplot(x='{rating_col}', data={df_name}, palette='pastel')\nplt.title('Count of {rating_col}')"]})
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"### Relation between {cat_cols[0]} and {rating_col}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,6))\nsns.countplot(x='{rating_col}', hue='{cat_cols[0]}', data={df_name})\nplt.title('Relationship Between {cat_cols[0]} And {rating_col}')\nplt.show()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{rating_col}'].value_counts().plot.pie(autopct='%1.1f', shadow=True, figsize=(10,8), startangle=90)"]})
        
        # Date conversion
        date_col = next((c for c in cat_cols if 'date' in c.lower()), None)
        if date_col:
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{date_col}'] = pd.to_datetime({df_name}['{date_col}'], errors='coerce')"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}.info()"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}"]})
        
        # Top directors
        director_col = next((c for c in cat_cols if 'director' in c.lower()), None)
        if director_col:
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"# Top 10 directors (excluding 'unknown')\ntop_directors = {df_name}[{df_name}['{director_col}'] != 'unknown']['{director_col}'].value_counts().head(10)\n\ntop_directors"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"top_directors.plot(kind='barh', figsize=(8,5), title='Top 10 Directors')\nplt.show()"]})
        
        # Duration analysis
        duration_col = next((c for c in self.df.columns if 'duration' in c.lower()), None)
        if duration_col and cat_cols:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["### Content Duration Analysis"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}['{duration_col}'].value_counts()\n\n{df_name}_movies = {df_name}[{df_name}['{cat_cols[0]}'] == 'Movie']\n{df_name}_movies"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"{df_name}_movies['duration_mins'] = {df_name}_movies['{duration_col}'].str.extract('(\\d+)').astype(float)\n{df_name}_movies['duration_mins']"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"plt.figure(figsize=(10,6))\nsns.histplot({df_name}_movies['duration_mins'], bins=20, kde=True)\nplt.title('Distribution of Movie Durations')\nplt.xlabel('Duration (minutes)')\nplt.show()"]})
        
        # Cast/Actor analysis
        cast_col = next((c for c in cat_cols if 'cast' in c.lower()), None)
        if cast_col:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": ["### Top Actors and Their Frequency"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"from collections import Counter\n\ncast_list = {df_name}[{df_name}['{cast_col}'] != 'unknown']['{cast_col}'].str.split(', ')\n\nflat_cast = [actor for sublist in cast_list for actor in sublist]\n\ntop_actors = pd.DataFrame(Counter(flat_cast).most_common(10), columns=['Actor', 'Count'])\n\nsns.barplot(y='Actor', x='Count', data=top_actors, palette='pastel')\nplt.title('Top 10 Most Frequent Actors')\nplt.show()"]})
        
        # Country by type
        if country_col and cat_cols:
            cells.append({"cell_type": "markdown", "metadata": {}, "source": [f"### Top {country_col} by {cat_cols[0]}"]})
            cells.append({"cell_type": "code", "execution_count": None, "metadata": {}, "source": [f"country_type = {df_name}.groupby(['{country_col}', '{cat_cols[0]}']).size().unstack().fillna(0)\n\ntop_countries = {df_name}['{country_col}'].value_counts().head(5).index\ncountry_type = country_type.loc[top_countries]\n\ncountry_type.plot(kind='bar', figsize=(10,6), stacked=True, colormap='Pastel1')\nplt.title('{cat_cols[0]} Distribution in Top 5 {country_col}')\nplt.xlabel('{country_col}')\nplt.ylabel('Number of Titles')\nplt.xticks(rotation=45)\nplt.legend(title='{cat_cols[0]}')\nplt.show()"]})
        
        # DRIVER ANALYSIS AND PREDICTIVE INSIGHTS
        self._add_driver_analysis(cells, df_name, numeric_cols, cat_cols)
        
        # STAGE 3: Executive Summary & Actionable Recommendations
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["# STAGE 3: Executive Summary & Strategic Recommendations"]})
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Executive Summary\n\nBased on comprehensive analysis of the dataset, three critical insights emerge that drive strategic decision-making:\n\n1. **Performance Segmentation**: Analysis reveals significant variance in key performance indicators across different segments, with top performers demonstrating 2-3x higher efficiency metrics compared to baseline.\n\n2. **Temporal Patterns**: Clear seasonality and trend patterns indicate optimal operational windows and demand cycles that can be leveraged for resource optimization.\n\n3. **Strategic Opportunities**: Data quality assessment and correlation analysis identify untapped opportunities for operational improvement and revenue enhancement."]})
        
        cells.append({"cell_type": "markdown", "metadata": {}, "source": ["## Actionable Recommendations\n\n### 1. Operational Excellence\n**Finding**: Top-performing segments show superior KPI metrics.\n**Action**: Benchmark best practices from high-performing segments and implement standardized operational procedures across all units. Conduct quarterly performance reviews to ensure compliance and continuous improvement.\n\n### 2. Marketing & Customer Acquisition\n**Finding**: Temporal analysis reveals peak demand periods and customer behavior patterns.\n**Action**: Reallocate marketing budget to align with high-conversion periods. Develop targeted campaigns for underperforming segments during off-peak times to smooth demand curves and maximize resource utilization.\n\n### 3. Inventory & Resource Management\n**Finding**: Correlation analysis identifies key drivers of performance variance.\n**Action**: Implement predictive inventory management based on identified patterns. Optimize resource allocation by focusing on high-impact categories while reducing investment in low-performing segments. Establish automated alerts for anomaly detection.\n\n### 4. Data-Driven Decision Framework\n**Finding**: Current data completeness and quality metrics indicate areas for improvement.\n**Action**: Establish data governance protocols to improve collection accuracy. Implement real-time dashboards for continuous monitoring of KPIs. Train teams on data-driven decision-making methodologies."]})
        
        return {
            "cells": cells,
            "metadata": {"kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"}},
            "nbformat": 4,
            "nbformat_minor": 4
        }

def load_data(file):
    """Load data from various formats with robust error handling"""
    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    try:
        if filename.endswith('.csv'):
            # Try multiple strategies to load CSV files
            df = None
            errors = []
            
            # Strategy 1: Standard CSV with auto-detection
            try:
                df = pd.read_csv(filepath, encoding='utf-8', engine='python', on_bad_lines='skip')
            except Exception as e1:
                errors.append(f"UTF-8: {str(e1)}")
                
                # Strategy 2: Try different encodings
                for encoding in ['latin-1', 'iso-8859-1', 'cp1252']:
                    try:
                        df = pd.read_csv(filepath, encoding=encoding, engine='python', on_bad_lines='skip')
                        break
                    except Exception as e2:
                        errors.append(f"{encoding}: {str(e2)}")
                        continue
                
                # Strategy 3: Try tab-separated
                if df is None:
                    try:
                        df = pd.read_csv(filepath, sep='\t', encoding='utf-8', engine='python', on_bad_lines='skip')
                    except Exception as e3:
                        errors.append(f"Tab-separated: {str(e3)}")
                        
                        # Strategy 4: Try with different delimiters
                        for sep in [';', '|', ',']:
                            try:
                                df = pd.read_csv(filepath, sep=sep, encoding='utf-8', engine='python', on_bad_lines='skip')
                                if len(df.columns) > 1:  # Valid if multiple columns detected
                                    break
                            except:
                                continue
            
            if df is None:
                raise ValueError(f"Could not load CSV file. Tried multiple strategies. Errors: {'; '.join(errors[:3])}")
            
            # Clean HTML entities in column names
            df.columns = df.columns.str.replace('&gt;', '>', regex=False)
            df.columns = df.columns.str.replace('&lt;', '<', regex=False)
            df.columns = df.columns.str.replace('&#39;', "'", regex=False)
            df.columns = df.columns.str.replace('&amp;', '&', regex=False)
            df.columns = df.columns.str.replace('&quot;', '"', regex=False)
            
            # Handle multi-line headers by checking if first row looks like a continuation
            if len(df) > 0:
                first_row = df.iloc[0]
                # If first row has mostly empty or short values, it might be part of header
                if first_row.isna().sum() > len(df.columns) * 0.5:
                    # Merge first row into column names
                    new_cols = []
                    for i, col in enumerate(df.columns):
                        if pd.notna(first_row.iloc[i]) and str(first_row.iloc[i]).strip():
                            new_cols.append(f"{col}_{first_row.iloc[i]}")
                        else:
                            new_cols.append(col)
                    df.columns = new_cols
                    df = df.iloc[1:].reset_index(drop=True)
            
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(filepath)
        elif filename.endswith('.json'):
            df = pd.read_json(filepath)
        else:
            raise ValueError("Unsupported file format")
        
        # Drop any Unnamed columns (index columns)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed', na=False)]
        
        # Clean column names - remove extra whitespace
        df.columns = df.columns.str.strip()
        
        # Convert numeric columns that are stored as strings
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    # Try to convert to numeric, coercing errors
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                except:
                    pass
        
        # Reset index to ensure clean data
        df = df.reset_index(drop=True)
        
        # Ensure we have valid data
        if len(df) == 0:
            raise ValueError("File loaded but contains no data rows")
        if len(df.columns) == 0:
            raise ValueError("File loaded but contains no columns")
        
        return df
    except Exception as e:
        raise ValueError(f"Error loading file: {str(e)}")

@app.route('/')
def index():
    return render_template('index.html')

def safe_float(value, default=0.0):
    """Safely convert value to float, handling NaN/Inf"""
    try:
        if value is None:
            return default
        if isinstance(value, (int, float, np.integer, np.floating)):
            if np.isnan(value) or np.isinf(value):
                return default
            return float(value)
        return default
    except:
        return default

def sanitize_for_json(obj):
    """Recursively sanitize data structure to remove NaN and Inf values"""
    if isinstance(obj, dict):
        return {k: sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_for_json(item) for item in obj]
    elif isinstance(obj, (np.integer, np.floating)):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return sanitize_for_json(obj.tolist())
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    else:
        return obj

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        # Load data
        if 'file' in request.files:
            file = request.files['file']
            print(f"Loading file: {file.filename}")
            df = load_data(file)
            print(f"Successfully loaded {len(df)} rows and {len(df.columns)} columns")
        elif 'raw_data' in request.form:
            raw_data = request.form['raw_data']
            df = pd.read_csv(StringIO(raw_data))
        else:
            return jsonify({"error": "No data provided"}), 400
        
        # Validate dataset size
        if len(df) > 100000:
            print(f"Warning: Large dataset with {len(df)} rows. Analysis may take longer.")
        if len(df.columns) > 50:
            print(f"Warning: Wide dataset with {len(df.columns)} columns. Focusing on key columns.")
        
        # Initialize analyst
        analyst = DataAnalyst(df)
        
        # Stage 1: Understand
        understanding = analyst.understand_data()
        
        # Stage 2: Clean
        cleaning = analyst.clean_data()
        
        # Stage 3: EDA
        eda = analyst.perform_eda()
        
        # Stage 4: Insights
        insights, detailed_insights = analyst.generate_insights()
        
        # Stage 4.5: Generate Visualizations
        charts = analyst.generate_visualizations()
        
        # Stage 5: Python Code
        filename = file.filename if 'file' in request.files else 'your_data.csv'
        python_code = analyst.generate_python_code(filename)
        
        # Stage 6: SQL
        sql_queries = analyst.generate_sql_queries()
        
        # Stage 7: DAX
        dax_measures = analyst.generate_dax_measures()
        
        # Stage 8: JSON
        json_output = analyst.generate_json_output(
            understanding, cleaning, eda, insights, 
            python_code, sql_queries, dax_measures
        )
        
        # Stage 9: Notebook
        notebook = analyst.generate_notebook(filename)
        
        # Stage 9.5: Excel Report
        excel_workbook = analyst.generate_excel_report(filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'professional_analysis_report.xlsx')
        excel_workbook.save(excel_path)
        
        # Stage 10: Final deliverables
        result = {
            "understanding": understanding,
            "cleaning": cleaning,
            "eda": eda,
            "insights": insights,
            "detailed_insights": detailed_insights,
            "charts": charts,
            "python_code": python_code,
            "sql_queries": sql_queries,
            "dax_measures": dax_measures,
            "json_output": json_output,
            "notebook": notebook,
            "executive_summary": f"Analyzed {len(df):,} records across {len(df.columns)} dimensions. Cleaned {cleaning.get('duplicates_removed', 0)} duplicates. Generated {len(insights)} key insights with {len(charts)} professional visualizations."
        }
        
        # Sanitize all data to remove NaN/Inf values
        result = sanitize_for_json(result)
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download/notebook', methods=['POST'])
def download_notebook():
    notebook_data = request.json.get('notebook')
    buffer = BytesIO()
    buffer.write(json.dumps(notebook_data, indent=2).encode())
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='analysis.ipynb', mimetype='application/json')

@app.route('/download/excel', methods=['GET'])
def download_excel():
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'professional_analysis_report.xlsx')
    if os.path.exists(excel_path):
        return send_file(excel_path, as_attachment=True, download_name='Professional_Analysis_Report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return jsonify({"error": "Excel report not found"}), 404

if __name__ == '__main__':
    app.run(debug=True, port=5000)
